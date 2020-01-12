import openpyxl
def writeToTxtFileFunction(rowNum,colNum,value):
    wb_obj = openpyxl.load_workbook("top100BooksByMedium.xlsx")
    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row
    print("rows=" + str(m_row))

    sheet_obj.cell(row=rowNum + 1, column=colNum).value = str(value) #col=4
    print("---rating write complete---")
    print("-----------------------")
    wb_obj.save("top100BooksByMedium.xlsx")


if __name__ == '__main__':
    print("unexpected calling")
