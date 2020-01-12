import openpyxl


def writeToExcelFunction(bookNestedDict):
    print("opening new excel for book author rating")

    wb_objex = openpyxl.Workbook()
    sheet_objex = wb_objex.active
    sheet_objex.cell(row=1, column=1).value = 'Sl. No'
    sheet_objex.cell(row=1, column=2).value = 'Book Name'
    sheet_objex.cell(row=1, column=3).value = 'Book Author'
    sheet_objex.cell(row=1, column=4).value = 'Rating'

    print("------writng list slno book author to excel start------")

    for bn, ba in bookNestedDict.items():
        print("book num", bn)

        col1 = bn  #booknum
        for key in ba:
            print(key + ":", ba[key])

            col2 = ba['name']  #bookname
            col3 = ba['author']  #bookauthor
        f = open("top100BokksByMedium.txt", "a")
        f.write(str(col1) + '\t' + str(bookname) + '\t' + bookauthor + '\n')
        f.close()
        sheet_objex.cell(row=col1 + 1, column=1).value = str(col1)
        sheet_objex.cell(row=col1 + 1, column=2).value = str(col2)
        sheet_objex.cell(row=col1 + 1, column=3).value = str(col3)
        # sheet_objex.cell(row=col1+1,column=4).value='Rating'
    print("------writng list slno book author to excel end------")

    wb_objex.save("top100BooksByMedium.xlsx")


if __name__ == '__main__':
    print("unexpected calling")
