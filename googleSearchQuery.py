import openpyxl
import googleSearch

def googleSearchQueryFunction():
    print("------saved excel file open------")
    wb_obj = openpyxl.load_workbook("top100BooksByMedium.xlsx")
    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row
    print("rows=" + str(m_row))

    for i in range(1, m_row):
        bookName = sheet_obj.cell(row=i + 1, column=2).value
        bookAuthor = sheet_obj.cell(row=i + 1, column=3).value
        searchPhrase = bookName + " by " + bookAuthor + " goodreads"
        # searchPhrase="Catch-22 by Joseph Heller in good reads"
        print("modified search phrase ", searchPhrase)
        googleSearch.googleSearchFunction(searchPhrase,i)

        #call google search and pass the searchPhrase


if __name__ == '__main__':
    print("unexpected calling")
